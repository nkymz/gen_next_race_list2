# -*- coding: utf-8 -*-

import os

import openpyxl

PPRO_PATH = os.getenv("HOMEDRIVE", "None") + os.getenv("HOMEPATH", "None") + "/Dropbox/POG/"
PO_HORSELIST_NAME = "POG_HorseList.xlsx"


class POHorseList:

    def __init__(self):
        wbpath = (PPRO_PATH + PO_HORSELIST_NAME).replace("\\", "/")
        self.wb = openpyxl.load_workbook(wbpath)
        self.ws = self.wb["POHorseList"]

    def get(self):
        trow = 1
        mylist = []

        while self.ws.cell(row=trow, column=1).value:

            horse_name = self.ws.cell(row=trow, column=2).value
            owner_name = self.ws.cell(row=trow, column=1).value
            horse_id = self.ws.cell(row=trow, column=7).value
            is_seal = True if self.ws.cell(row=trow, column=6).value == "-" else False
            mylist.append([horse_id, horse_name, owner_name, is_seal])

            trow += 1

        return mylist
